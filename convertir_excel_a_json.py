import json
from pathlib import Path
from openpyxl import load_workbook

EXCEL = "Πελάτες.xlsx"
OUT = "tiendas.json"

def clean_afm(x):
    if x is None:
        return ""
    s = str(x).strip()
    s = s.replace(".", "").replace(",", "").replace(" ", "")
    if s.endswith(".0"):
        s = s[:-2]
    return s

def norm(x):
    return "" if x is None else str(x).strip()

def main():
    if not Path(EXCEL).exists():
        raise FileNotFoundError(f"No encuentro el archivo: {EXCEL}")

    wb = load_workbook(EXCEL, data_only=True)
    ws = wb.active  # primera hoja

    tiendas = []

    # ESTRUCTURA REAL DEL EXCEL:
    # A = Κωδικός πελάτη (ΠΕΛ0000001) ❌
    # B = Επωνυμία ✅
    # C = Όνομα ✅
    # D = Κατηγορία
    # E = Α.Φ.Μ. REAL ✅

    for row in ws.iter_rows(min_row=2, values_only=True):
        afm = clean_afm(row[4])   # 🔴 COLUMNA E
        epwnymia = norm(row[1])   # B
        onoma = norm(row[2])      # C

        if not afm:
            continue

        tiendas.append({
            "afm": afm,
            "epwnymia": epwnymia,
            "onoma": onoma
        })

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(tiendas, f, ensure_ascii=False, indent=2)

    print(f"✅ Generado {OUT} con {len(tiendas)} tiendas")

if __name__ == "__main__":
    main()
